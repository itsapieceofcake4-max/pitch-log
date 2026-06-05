from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

def lerp(a, b, t):
    return int(round(a + (b - a) * t))

def depth_color(pct, stripe=False):
    # green RGB(66,140,51) -> red-brown RGB(180,80,40)
    t = pct / 100.0
    r = lerp(66, 180, t)
    g = lerp(140, 80, t)
    b = lerp(51, 40, t)
    if stripe:
        r = min(255, r + 12); g = min(255, g + 12); b = min(255, b + 12)
    return f"{r:02X}{g:02X}{b:02X}"

wb = Workbook()

# ── Sheet 1: depth map ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "敵陣深度マップ"
ws.sheet_view.showGridLines = False

center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# big heading
ws["A1"] = "Home 敵陣深度: 自陣ゴール 0% → 敵陣ゴール 100%（Away は左右ミラー）"
ws["A1"].font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1A3D1A")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:V1")
ws.row_dimensions[1].height = 26

ncols = 21           # 0%..100% step 5
nrows = 14
header_row = 3
first_data_row = 4

# direction labels
ws.cell(row=2, column=2, value="← 自陣（浅い）").font = Font(name=FONT, bold=True, color="2E7D32")
c = ws.cell(row=2, column=ncols + 1, value="敵陣（深い）→")
c.font = Font(name=FONT, bold=True, color="B45028")
c.alignment = Alignment(horizontal="right")

# column headers (depth %)
ws.cell(row=header_row, column=1, value="幅\\深度").alignment = center
ws.cell(row=header_row, column=1).font = Font(name=FONT, bold=True, size=9)
for j in range(ncols):
    pct = j * 5
    col = j + 2
    cell = ws.cell(row=header_row, column=col, value=f"{pct}%")
    cell.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0E1825")
    cell.alignment = center
    cell.border = border

# data grid
for i in range(nrows):
    r = first_data_row + i
    lbl = "上" if i == 0 else ("下" if i == nrows - 1 else "")
    lc = ws.cell(row=r, column=1, value=lbl)
    lc.font = Font(name=FONT, bold=True, size=9)
    lc.alignment = center
    for j in range(ncols):
        pct = j * 5
        col = j + 2
        zone_id = (j + 1) * 100 + i        # 列=深度バンド(×100), 行=幅(+1)
        cell = ws.cell(row=r, column=col, value=zone_id)
        cell.fill = PatternFill("solid", fgColor=depth_color(pct, stripe=(j % 2 == 1)))
        cell.font = Font(name=FONT, size=9, color="FFFFFF", bold=True)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[r].height = 22

# square-ish cells
ws.column_dimensions["A"].width = 9
for j in range(ncols):
    ws.column_dimensions[get_column_letter(j + 2)].width = 5.0

# note
note_r = first_data_row + nrows + 1
ws.cell(row=note_r, column=1,
        value="※ 値=ゾーンID（列の深度バンド×100 ＋ 行番号0〜13）。色は緑(自陣)→赤茶(敵陣)。列ヘッダー=Home敵陣深度%。")
ws.cell(row=note_r, column=1).font = Font(name=FONT, italic=True, size=9, color="555555")
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=ncols + 1)

# ── Sheet 2: coordinate mapping definition ─────────────────────────────────
ws2 = wb.create_sheet("座標マッピング定義")
ws2.sheet_view.showGridLines = False
hd_fill = PatternFill("solid", fgColor="1A3D1A")
hd_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)

ws2["A1"] = "Pitch Log — 座標マッピング定義（現状）"
ws2["A1"].font = Font(name=FONT, bold=True, size=13)
ws2.merge_cells("A1:B1")

rows = [
    ("項目", "定義"),
    ("原点 (0,0)", "ピッチ左下コーナー（正規化 0.0, 0.0）"),
    ("X 軸", "0(左ゴール) → 105m(右ゴール)、正規化 0.0→1.0"),
    ("Y 軸", "0(下) → 68m(上)、正規化 0.0→1.0"),
    ("グリッド", "105 × 68 セル（計 7140）"),
    ("GridID", "X_idx + Y_idx × 105 + 1（範囲 1〜7140）"),
    ("X_idx", "floor(x_norm × 105)（0〜104）"),
    ("Y_idx", "floor(y_norm × 68)（0〜67）"),
    ("Home 攻撃方向", "右（高xT=右ゴール付近）"),
    ("Away 攻撃方向", "左（ミラーマップ）"),
    ("Home 敵陣深度%", "x_norm × 100"),
    ("Away 敵陣深度%", "(1 − x_norm) × 100"),
]
start = 3
for idx, (a, b) in enumerate(rows):
    r = start + idx
    ca = ws2.cell(row=r, column=1, value=a)
    cb = ws2.cell(row=r, column=2, value=b)
    if idx == 0:
        ca.fill = hd_fill; cb.fill = hd_fill
        ca.font = hd_font; cb.font = hd_font
    else:
        ca.font = Font(name=FONT, bold=True, size=10)
        cb.font = Font(name=FONT, size=10)
    ca.alignment = Alignment(vertical="center")
    cb.alignment = Alignment(vertical="center", wrap_text=True)
    ca.border = border; cb.border = border

# GridID examples
ex_start = start + len(rows) + 1
ws2.cell(row=ex_start, column=1, value="GridID 実例").font = Font(name=FONT, bold=True, size=11)
examples = [
    ("左下 (0,0)", 1), ("右下 (104,0)", 105),
    ("左上 (0,67)", 7036), ("右上 (104,67)", 7140),
    ("中央 (52,34)", 3623),
]
for idx, (lbl, gid) in enumerate(examples):
    r = ex_start + 1 + idx
    ca = ws2.cell(row=r, column=1, value=lbl)
    cb = ws2.cell(row=r, column=2, value=gid)
    ca.font = Font(name=FONT, size=10); cb.font = Font(name=FONT, size=10)
    ca.border = border; cb.border = border

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 52

import sys
out = sys.argv[1] if len(sys.argv) > 1 else "pitch_depth_mapping.xlsx"
wb.save(out)
print("saved ->", out)
